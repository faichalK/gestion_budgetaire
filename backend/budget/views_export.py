"""
BudgetFlow — Exports Excel & PDF
  GET /api/budget/<pk>/export/budget-excel/
  GET /api/budget/<pk>/export/budget-pdf/
  GET /api/budget/<pk>/export/depenses-excel/
  GET /api/budget/<pk>/export/depenses-pdf/
"""
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework.views import APIView
from rest_framework import permissions, status
from rest_framework.response import Response
from datetime import date as dt_date

from .models import Budget, CategoriePrincipale, ConsommationLigne, StatutBudget


# ── Contrôle d'accès ──────────────────────────────────────────────────────────

def _peut_acceder(user, budget):
    if user.is_administrateur:
        return True
    if user.is_comptable and budget.statut != StatutBudget.BROUILLON:
        return True
    if user.is_gestionnaire and budget.gestionnaire_id == user.id:
        return True
    return False


# ── Helpers openpyxl ─────────────────────────────────────────────────────────

def _fill(hex_color):
    from openpyxl.styles import PatternFill
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')


def _border(style='thin'):
    from openpyxl.styles import Border, Side
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def _cell(ws, row, col, val=None, *, font=None, bg=None,
          align='left', num_fmt=None, wrap=False, valign='center'):
    from openpyxl.styles import Alignment
    c = ws.cell(row=row, column=col, value=val)
    c.border = _border()
    if font:    c.font = font
    if bg:      c.fill = _fill(bg)
    c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if num_fmt: c.number_format = num_fmt
    return c


def _fmt(n):
    try:
        return f"{float(n or 0):,.0f}"
    except (TypeError, ValueError):
        return "0"


# ══════════════════════════════════════════════════════════════════════════════
# BUDGET → EXCEL
# ══════════════════════════════════════════════════════════════════════════════

class ExportBudgetExcelView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        budget = get_object_or_404(
            Budget.objects.select_related('departement', 'gestionnaire'), pk=pk
        )
        if not _peut_acceder(request.user, budget):
            return Response({'detail': 'Accès non autorisé.'}, status=status.HTTP_403_FORBIDDEN)

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Lignes budgétaires"

        # ── Palette ──────────────────────────────────────────────────────────
        NAVY    = '1E3A8A'
        NAVY2   = '1D4ED8'
        NAVY_LT = 'DBEAFE'
        NAVY_XL = 'EFF6FF'
        GOLD    = 'F59E0B'
        GRAY50  = 'F9FAFB'
        GRAY100 = 'F3F4F6'
        WHITE   = 'FFFFFF'
        GREEN50 = 'F0FDF4'
        GREEN7  = '15803D'
        RED50   = 'FEF2F2'
        RED7    = 'B91C1C'
        AMB50   = 'FFFBEB'
        AMB7    = 'B45309'
        PURPLE5 = 'F5F3FF'
        PURPLE7 = '6D28D9'

        bold_white  = Font(bold=True, color=WHITE,  size=10)
        bold_navy   = Font(bold=True, color=NAVY,   size=10)
        bold_9      = Font(bold=True, size=9)
        normal_9    = Font(size=9)
        hdr_font    = Font(bold=True, color=WHITE,  size=9)
        sub_hdr     = Font(bold=True, color='374151', size=8)
        title_font  = Font(bold=True, color=NAVY,   size=16)
        info_key    = Font(bold=True, color='374151', size=9)
        info_val    = Font(color='111827', size=9)

        def kpi_box(row, col_start, col_end, label, value, bg, txt_color):
            ws.merge_cells(start_row=row,   start_column=col_start,
                           end_row=row,     end_column=col_end)
            ws.merge_cells(start_row=row+1, start_column=col_start,
                           end_row=row+1,   end_column=col_end)
            lc = ws.cell(row=row,   column=col_start, value=label)
            vc = ws.cell(row=row+1, column=col_start, value=value)
            for c in [lc, vc]:
                c.fill = _fill(bg)
                c.border = _border()
            lc.font      = Font(bold=True, color=txt_color, size=8)
            lc.alignment = Alignment(horizontal='center', vertical='bottom')
            vc.font      = Font(bold=True, color=txt_color, size=12)
            vc.alignment = Alignment(horizontal='center', vertical='top')
            # fill right-side cells of the merge
            for c_col in range(col_start + 1, col_end + 1):
                for r in [row, row + 1]:
                    x = ws.cell(row=r, column=c_col)
                    x.fill = _fill(bg)
                    x.border = _border()

        # ── Bannière titre ───────────────────────────────────────────────────
        ws.row_dimensions[1].height = 14
        ws.row_dimensions[2].height = 14

        ws.merge_cells('A1:F2')
        c = ws.cell(row=1, column=1, value=f"  BUDGET — {budget.nom}")
        c.font      = Font(bold=True, color=WHITE, size=14)
        c.fill      = _fill(NAVY)
        c.alignment = Alignment(horizontal='left', vertical='center')

        ws.merge_cells('G1:H2')
        d_exp = ws.cell(row=1, column=7,
                        value=f"Code : {budget.code}   |   Exporté le {dt_date.today().strftime('%d/%m/%Y')}")
        d_exp.font      = Font(color='BFDBFE', size=8)
        d_exp.fill      = _fill(NAVY)
        d_exp.alignment = Alignment(horizontal='right', vertical='center')

        # ── Infos ────────────────────────────────────────────────────────────
        row = 4
        gest = (f"{budget.gestionnaire.prenom} {budget.gestionnaire.nom}"
                if budget.gestionnaire else "—")
        infos = [
            ("Département",   budget.departement.nom if budget.departement else "—"),
            ("Période",       f"{budget.date_debut} → {budget.date_fin}"),
            ("Statut",        budget.get_statut_display()),
            ("Gestionnaire",  gest),
        ]
        for i, (label, val) in enumerate(infos):
            col_l = 1 + (i % 2) * 4
            col_v = col_l + 1
            if i % 2 == 0:
                cur_row = row + (i // 2)
            ws_row = row + (i // 2)
            _cell(ws, ws_row, col_l, label + " :", font=info_key, bg=GRAY50, align='right')
            ws.merge_cells(start_row=ws_row, start_column=col_v,
                           end_row=ws_row,   end_column=col_v + 1)
            _cell(ws, ws_row, col_v, val, font=info_val, bg=WHITE)
            ws.cell(row=ws_row, column=col_v + 1).fill = _fill(WHITE)
            ws.cell(row=ws_row, column=col_v + 1).border = _border()
        row = row + 2 + 1  # 2 info rows + 1 blank

        # ── KPI boxes ────────────────────────────────────────────────────────
        montant_global    = float(budget.montant_global    or 0)
        montant_consomme  = float(budget.montant_consomme  or 0)
        montant_disponible= float(budget.montant_disponible or 0)
        taux              = budget.calculer_taux_consommation()

        ws.row_dimensions[row].height   = 16
        ws.row_dimensions[row+1].height = 22

        kpi_box(row, 1, 2, "BUDGET GLOBAL",    f"{montant_global:,.0f} FCFA",    NAVY_LT, NAVY2)
        kpi_box(row, 3, 4, "CONSOMMÉ",         f"{montant_consomme:,.0f} FCFA",  PURPLE5, PURPLE7)
        kpi_box(row, 5, 6, "DISPONIBLE",
                f"{montant_disponible:,.0f} FCFA",
                GREEN50 if montant_disponible > 0 else RED50,
                GREEN7  if montant_disponible > 0 else RED7)
        kpi_box(row, 7, 8, "TAUX D'EXÉCUTION", f"{taux:.1f} %",
                AMB50 if taux < 75 else RED50,
                AMB7  if taux < 75 else RED7)
        row += 3  # 2 KPI rows + 1 blank

        # ── En-têtes colonnes ────────────────────────────────────────────────
        hdr_row = row
        headers = ["Code", "Désignation", "Unité", "Qté",
                   "Prix Unit. (FCFA)", "Alloué (FCFA)", "Consommé (FCFA)", "Disponible (FCFA)"]
        for col, h in enumerate(headers, 1):
            _cell(ws, hdr_row, col, h, font=hdr_font, bg=NAVY2, align='center')
        ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)
        ws.auto_filter.ref = f"A{hdr_row}:H{hdr_row}"
        row += 1

        # ── Données ──────────────────────────────────────────────────────────
        cats = CategoriePrincipale.objects.filter(budget=budget).prefetch_related(
            'sous_categories__lignes'
        ).order_by('ordre', 'code')

        grand_a = grand_c = grand_d = 0.0
        line_idx = 0  # pour les bandes alternées sur les lignes

        for cat in cats:
            scs   = list(cat.sous_categories.all())
            cat_a = sum(float(l.montant_alloue   or 0) for sc in scs for l in sc.lignes.all())
            cat_c = sum(float(l.montant_consomme or 0) for sc in scs for l in sc.lignes.all())
            cat_d = cat_a - cat_c

            # ── Catégorie ──
            ws.merge_cells(f'A{row}:E{row}')
            _cell(ws, row, 1, f"  {cat.code}  —  {cat.libelle}", font=bold_white, bg=NAVY)
            for col in [2, 3, 4, 5]:
                ws.cell(row=row, column=col).fill   = _fill(NAVY)
                ws.cell(row=row, column=col).border = _border()
            _cell(ws, row, 6, cat_a, font=bold_white, bg=NAVY, align='right', num_fmt='#,##0')
            _cell(ws, row, 7, cat_c, font=bold_white, bg=NAVY, align='right', num_fmt='#,##0')
            _cell(ws, row, 8, cat_d, font=bold_white, bg=NAVY, align='right', num_fmt='#,##0')
            ws.row_dimensions[row].height = 18
            row += 1

            for sc in scs:
                lignes = list(sc.lignes.all())
                sc_a   = sum(float(l.montant_alloue   or 0) for l in lignes)
                sc_c   = sum(float(l.montant_consomme or 0) for l in lignes)
                sc_d   = sc_a - sc_c

                # ── Sous-catégorie ──
                ws.merge_cells(f'A{row}:E{row}')
                _cell(ws, row, 1, f"    {sc.code}  {sc.libelle}",
                      font=Font(bold=True, color=NAVY2, size=9), bg=NAVY_XL)
                for col in [2, 3, 4, 5]:
                    ws.cell(row=row, column=col).fill   = _fill(NAVY_XL)
                    ws.cell(row=row, column=col).border = _border()
                _cell(ws, row, 6, sc_a, font=Font(bold=True, color=NAVY2, size=9),
                      bg=NAVY_XL, align='right', num_fmt='#,##0')
                _cell(ws, row, 7, sc_c, font=Font(bold=True, color=NAVY2, size=9),
                      bg=NAVY_XL, align='right', num_fmt='#,##0')
                _cell(ws, row, 8, sc_d, font=Font(bold=True, color=NAVY2, size=9),
                      bg=NAVY_XL, align='right', num_fmt='#,##0')
                row += 1

                if lignes:
                    # Sous-en-têtes de colonnes
                    sub_hdrs = ["", "Désignation", "Unité", "Qté", "Prix Unit.",
                                "Alloué", "Consommé", "Disponible"]
                    for col, h in enumerate(sub_hdrs, 1):
                        _cell(ws, row, col, h, font=sub_hdr, bg=GRAY50, align='center')
                    row += 1

                    for ligne in lignes:
                        alt_bg = WHITE if line_idx % 2 == 0 else GRAY100
                        line_idx += 1
                        _cell(ws, row, 1, f"      {ligne.code}",          font=normal_9, bg=alt_bg)
                        _cell(ws, row, 2, ligne.libelle,                   font=normal_9, bg=alt_bg, wrap=True)
                        _cell(ws, row, 3, ligne.unite or "—",              font=normal_9, bg=alt_bg, align='center')
                        _cell(ws, row, 4, float(ligne.quantite or 0),     font=normal_9, bg=alt_bg, align='right', num_fmt='#,##0.##')
                        _cell(ws, row, 5, float(ligne.prix_unitaire or 0),font=normal_9, bg=alt_bg, align='right', num_fmt='#,##0')
                        _cell(ws, row, 6, float(ligne.montant_alloue   or 0), font=normal_9, bg=alt_bg, align='right', num_fmt='#,##0')
                        _cell(ws, row, 7, float(ligne.montant_consomme or 0), font=normal_9, bg=alt_bg, align='right', num_fmt='#,##0')

                        dispo = float(ligne.montant_disponible or 0)
                        dispo_bg  = GREEN50 if dispo >= 0 else RED50
                        dispo_fnt = Font(color=GREEN7 if dispo >= 0 else RED7, size=9, bold=True)
                        _cell(ws, row, 8, dispo, font=dispo_fnt, bg=dispo_bg, align='right', num_fmt='#,##0')
                        row += 1

            grand_a += cat_a
            grand_c += cat_c
            grand_d += cat_d

        # ── Total général ──
        ws.merge_cells(f'A{row}:E{row}')
        _cell(ws, row, 1, "  TOTAL GÉNÉRAL", font=bold_white, bg=NAVY)
        for col in [2, 3, 4, 5]:
            ws.cell(row=row, column=col).fill   = _fill(NAVY)
            ws.cell(row=row, column=col).border = _border()
        _cell(ws, row, 6, grand_a, font=bold_white, bg=NAVY, align='right', num_fmt='#,##0')
        _cell(ws, row, 7, grand_c, font=bold_white, bg=NAVY, align='right', num_fmt='#,##0')
        grand_dispo_bg  = GREEN50 if grand_d >= 0 else RED50
        grand_dispo_fnt = Font(bold=True, color=GREEN7 if grand_d >= 0 else RED7, size=10)
        _cell(ws, row, 8, grand_d, font=grand_dispo_fnt, bg=grand_dispo_bg, align='right', num_fmt='#,##0')
        ws.row_dimensions[row].height = 20

        # ── Largeurs colonnes ──
        for i, w in enumerate([10, 40, 10, 8, 16, 16, 16, 16], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        resp = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        safe = budget.code.replace('/', '_')
        resp['Content-Disposition'] = f'attachment; filename="Budget_{safe}.xlsx"'
        wb.save(resp)
        return resp


# ══════════════════════════════════════════════════════════════════════════════
# BUDGET → PDF
# ══════════════════════════════════════════════════════════════════════════════

class ExportBudgetPdfView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        budget = get_object_or_404(
            Budget.objects.select_related('departement', 'gestionnaire'), pk=pk
        )
        if not _peut_acceder(request.user, budget):
            return Response({'detail': 'Accès non autorisé.'}, status=status.HTTP_403_FORBIDDEN)

        from exports.pdf.generators.budget_detail import generate_budget_detail_pdf

        user_name = f"{request.user.prenom} {request.user.nom}".strip() or request.user.email
        doc, elements = generate_budget_detail_pdf(budget, user_name)

        filename = doc.make_filename('budget', budget.code)
        return doc.as_response(elements, filename)


# ══════════════════════════════════════════════════════════════════════════════
# DÉPENSES → EXCEL
# ══════════════════════════════════════════════════════════════════════════════

class ExportDepensesExcelView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        budget = get_object_or_404(
            Budget.objects.select_related('departement', 'gestionnaire'), pk=pk
        )
        if not _peut_acceder(request.user, budget):
            return Response({'detail': 'Accès non autorisé.'}, status=status.HTTP_403_FORBIDDEN)

        depenses = ConsommationLigne.objects.filter(
            ligne__budget=budget
        ).select_related(
            'ligne__sous_categorie__categorie', 'enregistre_par', 'depense'
        ).order_by(
            'ligne__sous_categorie__categorie__ordre',
            'ligne__sous_categorie__categorie__code',
            'ligne__sous_categorie__ordre',
            'ligne__code', '-date',
        )

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter

        wb  = Workbook()
        ws  = wb.active
        ws.title = "Dépenses"

        NAVY    = '1E3A8A'
        NAVY2   = '1D4ED8'
        NAVY_LT = 'DBEAFE'
        GRAY50  = 'F9FAFB'
        GRAY100 = 'F3F4F6'
        WHITE   = 'FFFFFF'
        C_WAIT  = 'FEF3C7'; C_WAIT_T = 'B45309'
        C_OK    = 'DCFCE7'; C_OK_T   = '15803D'
        C_REJ   = 'FEE2E2'; C_REJ_T  = 'B91C1C'

        bold_white  = Font(bold=True, color=WHITE, size=10)
        bold_9      = Font(bold=True, size=9)
        normal_9    = Font(size=9)
        hdr_font    = Font(bold=True, color=WHITE, size=9)

        STATUT_LABELS = {'SAISIE': 'En attente', 'VALIDEE': 'Validée',  'REJETEE': 'Rejetée'}
        STATUT_BG     = {'SAISIE': C_WAIT,        'VALIDEE': C_OK,       'REJETEE': C_REJ}
        STATUT_FG     = {'SAISIE': C_WAIT_T,      'VALIDEE': C_OK_T,     'REJETEE': C_REJ_T}

        dep_list  = list(depenses)
        nb_valid  = sum(1 for d in dep_list if d.statut == 'VALIDEE')
        nb_wait   = sum(1 for d in dep_list if d.statut == 'SAISIE')
        nb_rej    = sum(1 for d in dep_list if d.statut == 'REJETEE')
        total_val = sum(float(d.montant or 0) for d in dep_list if d.statut == 'VALIDEE')
        grand_tot = sum(float(d.montant or 0) for d in dep_list)

        # ── Bannière titre ──
        ws.merge_cells('A1:H2')
        c = ws.cell(row=1, column=1, value=f"  DÉPENSES — {budget.nom}")
        c.font      = Font(bold=True, color=WHITE, size=14)
        c.fill      = _fill(NAVY)
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 14
        ws.row_dimensions[2].height = 14
        for col in range(2, 9):
            for r in [1, 2]:
                x = ws.cell(row=r, column=col)
                x.fill = _fill(NAVY); x.border = _border()
        d_exp = ws.cell(row=1, column=7,
                        value=f"Code : {budget.code}   |   Exporté le {dt_date.today().strftime('%d/%m/%Y')}")
        d_exp.font = Font(color='BFDBFE', size=8)
        d_exp.fill = _fill(NAVY)
        d_exp.alignment = Alignment(horizontal='right', vertical='center')
        ws.merge_cells('G1:H2')
        ws.cell(row=2, column=7).fill = _fill(NAVY)

        # ── Infos ──
        row = 4
        gest = f"{budget.gestionnaire.prenom} {budget.gestionnaire.nom}" if budget.gestionnaire else "—"
        for label, val in [("Département :", budget.departement.nom if budget.departement else "—"),
                            ("Gestionnaire :", gest)]:
            _cell(ws, row, 1, label, font=bold_9, bg=GRAY50, align='right')
            ws.merge_cells(f'B{row}:H{row}')
            _cell(ws, row, 2, val, font=normal_9)
            ws.cell(row=row, column=2).border = _border()
            row += 1
        row += 1

        # ── KPI résumé ──
        ws.row_dimensions[row].height   = 16
        ws.row_dimensions[row+1].height = 22

        def kpi_box(r, cs, ce, label, value, bg, tc):
            ws.merge_cells(start_row=r,   start_column=cs, end_row=r,   end_column=ce)
            ws.merge_cells(start_row=r+1, start_column=cs, end_row=r+1, end_column=ce)
            lc = ws.cell(row=r,   column=cs, value=label)
            vc = ws.cell(row=r+1, column=cs, value=value)
            for c in [lc, vc]:
                c.fill = _fill(bg); c.border = _border()
            lc.font = Font(bold=True, color=tc, size=8)
            lc.alignment = Alignment(horizontal='center', vertical='bottom')
            vc.font = Font(bold=True, color=tc, size=12)
            vc.alignment = Alignment(horizontal='center', vertical='top')
            for c_col in range(cs + 1, ce + 1):
                for rr in [r, r + 1]:
                    x = ws.cell(row=rr, column=c_col)
                    x.fill = _fill(bg); x.border = _border()

        kpi_box(row, 1, 2, "TOTAL DÉPENSES",   f"{grand_tot:,.0f} FCFA", NAVY_LT, NAVY2)
        kpi_box(row, 3, 4, "VALIDÉES",          f"{nb_valid}  ·  {total_val:,.0f} FCFA", C_OK, C_OK_T)
        kpi_box(row, 5, 6, "EN ATTENTE",        f"{nb_wait} dépense(s)", C_WAIT, C_WAIT_T)
        kpi_box(row, 7, 8, "REJETÉES",          f"{nb_rej} dépense(s)",  C_REJ,  C_REJ_T)
        row += 3

        # ── En-têtes colonnes ──
        hdr_row = row
        headers = ["Catégorie", "Sous-catégorie", "Ligne budgétaire",
                   "Référence", "Fournisseur", "Montant (FCFA)", "Date", "Statut"]
        for col, h in enumerate(headers, 1):
            _cell(ws, hdr_row, col, h, font=hdr_font, bg=NAVY2, align='center')
        ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)
        ws.auto_filter.ref = f"A{hdr_row}:H{hdr_row}"
        row += 1

        for line_idx, d in enumerate(dep_list):
            sc  = d.ligne.sous_categorie if (d.ligne and d.ligne.sous_categorie_id) else None
            cat = sc.categorie if (sc and sc.categorie_id) else None

            montant = float(d.montant or 0)
            bg      = STATUT_BG.get(d.statut, WHITE)
            fg      = STATUT_FG.get(d.statut, '111827')

            _cell(ws, row, 1, f"{cat.code} — {cat.libelle}" if cat else "—",
                  font=normal_9, bg=bg)
            _cell(ws, row, 2, f"{sc.code} — {sc.libelle}"   if sc  else "—",
                  font=normal_9, bg=bg)
            _cell(ws, row, 3, d.ligne.libelle if d.ligne else "—", font=normal_9, bg=bg)
            _cell(ws, row, 4, d.reference or str(d.id)[:8].upper(),
                  font=normal_9, bg=bg, align='center')
            _cell(ws, row, 5, d.fournisseur or (d.depense.fournisseur if d.depense_id else '') or "—", font=normal_9, bg=bg)
            _cell(ws, row, 6, montant, font=Font(bold=True, size=9), bg=bg,
                  align='right', num_fmt='#,##0')
            _cell(ws, row, 7, d.date.strftime('%d/%m/%Y') if d.date else "—",
                  font=normal_9, bg=bg, align='center')
            statut_label = STATUT_LABELS.get(d.statut, d.statut)
            _cell(ws, row, 8, statut_label,
                  font=Font(bold=True, color=fg, size=9), bg=bg, align='center')
            row += 1

        # ── Total général ──
        ws.merge_cells(f'A{row}:E{row}')
        _cell(ws, row, 1, "  TOTAL GÉNÉRAL", font=bold_white, bg=NAVY, align='left')
        for col in [2, 3, 4, 5]:
            ws.cell(row=row, column=col).fill = _fill(NAVY)
            ws.cell(row=row, column=col).border = _border()
        _cell(ws, row, 6, grand_tot, font=bold_white, bg=NAVY, align='right', num_fmt='#,##0')
        for col in [7, 8]:
            ws.cell(row=row, column=col).fill = _fill(NAVY)
            ws.cell(row=row, column=col).border = _border()
        ws.row_dimensions[row].height = 20

        for i, w in enumerate([22, 22, 28, 14, 22, 16, 14, 14], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        resp = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        safe = budget.code.replace('/', '_')
        resp['Content-Disposition'] = f'attachment; filename="Depenses_{safe}.xlsx"'
        wb.save(resp)
        return resp


# ══════════════════════════════════════════════════════════════════════════════
# DÉPENSES → PDF
# ══════════════════════════════════════════════════════════════════════════════

class ExportDepensesPdfView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        budget = get_object_or_404(
            Budget.objects.select_related('departement', 'gestionnaire'), pk=pk
        )
        if not _peut_acceder(request.user, budget):
            return Response({'detail': 'Accès non autorisé.'}, status=status.HTTP_403_FORBIDDEN)

        depenses = ConsommationLigne.objects.filter(
            ligne__budget=budget
        ).select_related(
            'ligne__sous_categorie__categorie', 'enregistre_par', 'depense'
        ).order_by(
            'ligne__sous_categorie__categorie__ordre',
            'ligne__sous_categorie__categorie__code',
            'ligne__sous_categorie__ordre',
            'ligne__code', '-date',
        )

        from exports.pdf.generators.depense_list import generate_depense_list_pdf

        user_name = f"{request.user.prenom} {request.user.nom}".strip() or request.user.email
        doc, elements = generate_depense_list_pdf(budget, depenses, user_name)

        filename = doc.make_filename('depenses', budget.code)
        return doc.as_response(elements, filename)
